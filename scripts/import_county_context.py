"""Import Illinois county disparity context and aggregate against matched facility data."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


COUNTY_HEALTH_RANKINGS_URL = "https://www.countyhealthrankings.org/health-data/illinois/data-and-resources"


def load_openpyxl():
    try:
        from openpyxl import load_workbook

        return load_workbook
    except ImportError:
        bundled = Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime" / "dependencies" / "python"
        if bundled.exists():
            sys.path.insert(0, str(bundled))
            from openpyxl import load_workbook

            return load_workbook
        raise


def clean_county(value: object) -> str:
    return str(value or "").replace(" County", "").strip().upper()


def number_or_none(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_sheet_by_county(workbook, sheet_name: str) -> dict[str, dict[str, object]]:
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[1]]
    records: dict[str, dict[str, object]] = {}

    for row in rows[3:]:
        values = dict(zip(headers, row))
        county = clean_county(values.get("County"))
        if not county:
            continue
        records[county] = values

    return records


def rural_classification(percent_rural: float | None) -> str:
    if percent_rural is None:
        return "Unknown"
    if percent_rural >= 0.5:
        return "Rural"
    if percent_rural >= 0.2:
        return "Mixed"
    return "Urban"


def load_county_context(workbook_path: Path) -> list[dict[str, object]]:
    load_workbook = load_openpyxl()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    selected = read_sheet_by_county(workbook, "Select Measure Data")
    additional = read_sheet_by_county(workbook, "Additional Measure Data")
    counties = sorted(set(selected) | set(additional))
    output: list[dict[str, object]] = []

    for county in counties:
        selected_row = selected.get(county, {})
        additional_row = additional.get(county, {})
        percent_rural = number_or_none(additional_row.get("% Rural"))
        output.append({
            "county": county.title(),
            "countyKey": county,
            "fips": str(selected_row.get("FIPS") or additional_row.get("FIPS") or ""),
            "ruralUrbanClassification": rural_classification(percent_rural),
            "percentRural": percent_rural,
            "medianHouseholdIncome": number_or_none(additional_row.get("Median Household Income")),
            "povertyRate": number_or_none(selected_row.get("% Children in Poverty")),
            "povertyMeasure": "% Children in Poverty",
            "age65PlusPercent": number_or_none(additional_row.get("% 65 and Over")),
            "uninsuredRate": number_or_none(selected_row.get("% Uninsured")),
            "population": number_or_none(additional_row.get("Population") or selected_row.get("Population")),
            "hospitalAccessIndicator": number_or_none(selected_row.get("Preventable Hospitalization Rate")),
            "primaryCarePhysiciansRate": number_or_none(selected_row.get("Primary Care Physicians Rate")),
            "source": "County Health Rankings & Roadmaps 2025 Illinois Data",
            "sourceUrl": COUNTY_HEALTH_RANKINGS_URL,
        })

    return output


def average(values: list[float]) -> float | None:
    filtered = [value for value in values if value is not None]
    if not filtered:
        return None
    return sum(filtered) / len(filtered)


def aggregate_facility_context(
    county_context: list[dict[str, object]],
    matched_facilities: list[dict[str, object]],
) -> list[dict[str, object]]:
    context_by_county = {record["countyKey"]: record for record in county_context}
    grouped: dict[str, list[dict[str, object]]] = {}

    for facility in matched_facilities:
        county = clean_county(facility.get("quality", {}).get("county"))
        if county:
            grouped.setdefault(county, []).append(facility)

    summaries: list[dict[str, object]] = []
    for county, facilities in sorted(grouped.items()):
        context = context_by_county.get(county, {
            "county": county.title(),
            "countyKey": county,
            "ruralUrbanClassification": "Unknown",
        })
        summary = dict(context)
        summary.update({
            "matchedFacilityCount": len(facilities),
            "averageTotalRate": average([facility.get("publishedAmount") for facility in facilities]),
            "averageCapitalRate": average([facility.get("components", {}).get("capitalRate") for facility in facilities]),
            "averageOverallStarRating": average([facility.get("quality", {}).get("overallStarRating") for facility in facilities]),
            "averageStaffingRating": average([facility.get("quality", {}).get("staffingStarRating") for facility in facilities]),
        })
        summaries.append(summary)

    return summaries


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("county_health_rankings_xlsx", type=Path)
    parser.add_argument("matched_quality_json", type=Path)
    parser.add_argument("county_context_output", type=Path)
    parser.add_argument("county_summary_output", type=Path)
    args = parser.parse_args()

    county_context = load_county_context(args.county_health_rankings_xlsx)
    matched_facilities = json.loads(args.matched_quality_json.read_text(encoding="utf-8"))
    county_summary = aggregate_facility_context(county_context, matched_facilities)

    args.county_context_output.parent.mkdir(parents=True, exist_ok=True)
    args.county_context_output.write_text(json.dumps(county_context, indent=2), encoding="utf-8")
    args.county_summary_output.write_text(json.dumps(county_summary, indent=2), encoding="utf-8")

    print(f"Wrote {len(county_context)} county context rows and {len(county_summary)} county summaries.")


if __name__ == "__main__":
    main()
