"""Import Illinois HFS nursing facility rate XLSX files into dashboard JSON."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path


SOURCE_NAME = "Illinois HFS Medicaid Rate List for Nursing Facilities"
SOURCE_URL = (
    "https://hfs.illinois.gov/medicalproviders/medicaidreimbursement/ltc/"
    "archivedmedicaidratelists.html"
)

SHEET_CATEGORIES = {
    "SNF": "Nursing Facility",
    "CEA": "County Nursing Facility",
    "SMHRF": "Specialized Mental Health Rehabilitation Facility",
}

HSA_REGIONS = {
    1: {
        "name": "HSA I",
        "description": "Northwest Illinois",
        "tier": "Downstate / Smaller Market",
    },
    2: {
        "name": "HSA II",
        "description": "Peoria and North-Central Illinois",
        "tier": "Regional Urban / Mixed",
    },
    3: {
        "name": "HSA III",
        "description": "West-Central Illinois",
        "tier": "Downstate / Smaller Market",
    },
    4: {
        "name": "HSA IV",
        "description": "Champaign, Bloomington, Decatur, and East-Central Illinois",
        "tier": "Regional Urban / Mixed",
    },
    5: {
        "name": "HSA V",
        "description": "Southern Illinois",
        "tier": "Downstate / Smaller Market",
    },
    6: {
        "name": "HSA VI",
        "description": "City of Chicago",
        "tier": "Chicago Metro",
    },
    7: {
        "name": "HSA VII",
        "description": "DuPage County and suburban Cook County",
        "tier": "Chicago Metro",
    },
    8: {
        "name": "HSA VIII",
        "description": "Kane, Lake, and McHenry Counties",
        "tier": "Chicago Metro",
    },
    9: {
        "name": "HSA IX",
        "description": "Grundy, Kankakee, Kendall, and Will Counties",
        "tier": "Chicago Metro",
    },
    10: {
        "name": "HSA X",
        "description": "Henry, Mercer, and Rock Island Counties",
        "tier": "Downstate / Smaller Market",
    },
    11: {
        "name": "HSA XI",
        "description": "Metro East Illinois",
        "tier": "Regional Urban / Mixed",
    },
}


def load_openpyxl():
    try:
        from openpyxl import load_workbook

        return load_workbook
    except ImportError:
        bundled = Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime" / "dependencies" / "python"
        if bundled.exists():
            sys.path.insert(0, str(bundled))
            from openpyxl import load_workbook

            return load_workbook
        raise


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def normalize_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, ""):
        return None
    return str(value)


def normalize_number(value: object) -> float | None:
    if value in (None, ""):
        return None
    return round(float(value), 2)


def normalize_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def category_for_sheet(sheet_name: str) -> str:
    for key, category in SHEET_CATEGORIES.items():
        if key in sheet_name.upper():
            return category
    return "Long-Term Care Facility"


def import_rates(input_path: Path) -> list[dict[str, object]]:
    load_workbook = load_openpyxl()
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    records: list[dict[str, object]] = []

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        title = next(rows, ())
        next(rows, ())
        subtitle = next(rows, ())
        headers = [normalize_header(value) for value in next(rows, ())]
        category = category_for_sheet(sheet.title)
        report_title = " ".join(str(value) for value in [title[0] if title else "", subtitle[0] if subtitle else ""] if value)

        for row in rows:
            values = dict(zip(headers, row))
            building_id = values.get("building_id")
            facility_name = values.get("facility_name")
            if not building_id or not facility_name:
                continue

            total_rate = normalize_number(values.get("total_rate"))
            rate_area = normalize_int(values.get("rate_area"))
            hsa = normalize_int(values.get("hsa")) or rate_area
            region = HSA_REGIONS.get(hsa, {})
            record = {
                "facility": str(facility_name).strip(),
                "city": str(values.get("city") or "").strip(),
                "category": category,
                "payer": "Illinois Medicaid",
                "service": "Long-term care facility per diem",
                "codeType": "HFS Building ID",
                "code": str(building_id).strip(),
                "publishedAmount": total_rate,
                "amountLabel": "Total rate",
                "effectiveDate": normalize_date(values.get("effective_rate")),
                "source": SOURCE_NAME,
                "confidence": "source-imported",
                "notes": (
                    f"{report_title}. Capital: {normalize_number(values.get('capital_rate'))}; "
                    f"support: {normalize_number(values.get('support_rate'))}; "
                    f"nursing: {normalize_number(values.get('nursing_rate'))}; "
                    f"HSA: {hsa}; rate area: {rate_area}."
                ),
                "components": {
                    "capitalRate": normalize_number(values.get("capital_rate")),
                    "supportRate": normalize_number(values.get("support_rate")),
                    "nursingRate": normalize_number(values.get("nursing_rate")),
                    "hsa": hsa,
                    "rateArea": rate_area,
                    "capitalRateChangeEffectiveDate": normalize_date(values.get("capital_rate_change_eff_date")),
                },
                "geography": {
                    "hsa": hsa,
                    "hsaName": region.get("name", f"HSA {hsa}" if hsa else "Unknown HSA"),
                    "region": region.get("description", "Unknown HSA region"),
                    "tier": region.get("tier", "Unclassified"),
                    "classificationBasis": "HSA-based proxy from Illinois facility rate file",
                },
                "sourceUrl": SOURCE_URL,
            }
            records.append(record)

    return records


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Path to the Illinois HFS nursing facility XLSX file.")
    parser.add_argument("output", type=Path, help="Path for dashboard-ready JSON output.")
    args = parser.parse_args()

    records = import_rates(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(records, indent=2), encoding="utf-8")
    print(f"Wrote {len(records)} records to {args.output}")


if __name__ == "__main__":
    main()
